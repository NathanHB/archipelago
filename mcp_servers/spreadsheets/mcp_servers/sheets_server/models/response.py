import re
from typing import Any

from mcp_schema import FlatBaseModel as BaseModel
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import ConfigDict


def _format_table_output(
    num_cols: int,
    values: list[list[Any]],
    headers: list[str] | None = None,
    formulas: dict[str, str] | None = None,
) -> str:
    """Format a table output with column headers, optional data headers, and values.

    Args:
        num_cols: Number of columns in the table
        values: List of data rows
        headers: Optional list of header strings (shown as "H" row)
        formulas: Optional dict mapping cell references (e.g., "A2") to formula strings

    Returns:
        Formatted table string with tab-separated columns
    """
    lines = []

    # Column header line (A, B, C, ...)
    header_line = "\t" + "\t".join(get_column_letter(i + 1) for i in range(num_cols))
    lines.append(header_line)

    start_row = 1
    # Add headers row if present
    if headers:
        row_data = ["H"]
        for col_idx in range(num_cols):
            if col_idx < len(headers):
                cell_str = str(headers[col_idx]) if headers[col_idx] is not None else ""
            else:
                cell_str = ""
            row_data.append(cell_str)
        lines.append("\t".join(row_data))
        start_row = 2

    # Add data rows
    for row_idx, row in enumerate(values):
        actual_row_num = start_row + row_idx
        row_data = [str(actual_row_num)]
        for col_idx in range(num_cols):
            if col_idx < len(row):
                cell_value = row[col_idx]
                if formulas:
                    actual_col_idx = col_idx + 1
                    cell_ref = f"{get_column_letter(actual_col_idx)}{actual_row_num}"
                    if cell_ref in formulas:
                        formula = formulas[cell_ref]
                        if cell_value is None or cell_value == "":
                            cell_str = f"({formula})"
                        else:
                            cell_str = f"{cell_value} ({formula})"
                    else:
                        cell_str = str(cell_value) if cell_value is not None else ""
                else:
                    cell_str = str(cell_value) if cell_value is not None else ""
            else:
                cell_str = ""
            row_data.append(cell_str)
        lines.append("\t".join(row_data))

    return "\n".join(lines)


class ReadTabSingleCellResponse(BaseModel):
    """Response for reading a single cell."""

    model_config = ConfigDict(extra="forbid")

    cell: str
    value: Any
    formula: str | None = None

    def __str__(self) -> str:
        base = f"{{'cell': '{self.cell}', 'value': {repr(self.value)}"
        if self.formula is not None:
            base += f", 'formula': {repr(self.formula)}"
        base += "}"
        return base


class ReadTabRangeResponse(BaseModel):
    """Response for reading a cell range or entire sheet."""

    model_config = ConfigDict(extra="forbid")

    range: str
    values: list[list[Any]]
    formulas: dict[str, str] | None = None

    def __str__(self) -> str:
        if not self.values:
            return f"Range: {self.range}\nTable: (empty)"

        num_cols = max(len(row) for row in self.values) if self.values else 0

        start_col_idx = 1
        start_row_idx = 1
        if self.range != "all":
            match = re.match(r"([A-Z]+)(\d+)", self.range.split(":")[0])
            if match:
                start_col_idx = column_index_from_string(match.group(1))
                start_row_idx = int(match.group(2))

        lines = []

        header = "\t" + "\t".join(
            get_column_letter(start_col_idx + i) for i in range(num_cols)
        )
        lines.append(header)

        for row_idx, row in enumerate(self.values):
            actual_row_num = start_row_idx + row_idx
            row_data = [str(actual_row_num)]
            for col_idx in range(num_cols):
                if col_idx < len(row):
                    cell_value = row[col_idx]
                    actual_col_idx = start_col_idx + col_idx
                    cell_ref = f"{get_column_letter(actual_col_idx)}{actual_row_num}"

                    if self.formulas and cell_ref in self.formulas:
                        formula = self.formulas[cell_ref]
                        if cell_value is None or cell_value == "":
                            cell_str = f"({formula})"
                        else:
                            cell_str = f"{cell_value} ({formula})"
                    else:
                        cell_str = str(cell_value) if cell_value is not None else ""
                else:
                    cell_str = ""

                row_data.append(cell_str)

            lines.append("\t".join(row_data))

        table = "\n".join(lines)
        return f"Range: {self.range}\nTable:\n{table}"


class WorksheetInfo(BaseModel):
    """Information about a worksheet tab."""

    model_config = ConfigDict(extra="forbid")

    name: str
    index: int
    row_count: int
    column_count: int


class ListTabsResponse(BaseModel):
    """Response for listing worksheet tabs in a spreadsheet."""

    model_config = ConfigDict(extra="forbid")

    worksheets: list[WorksheetInfo]

    def __str__(self) -> str:
        worksheets_str = ", ".join(
            f"{{'name': '{ws.name}', 'index': {ws.index}, 'row_count': {ws.row_count}, 'column_count': {ws.column_count}}}"
            for ws in self.worksheets
        )
        return f"{{'worksheets': [{worksheets_str}]}}"


class CreateSpreadsheetResponse(BaseModel):
    """Response for creating a spreadsheet."""

    model_config = ConfigDict(extra="forbid")

    status: str
    file_name: str
    file_path: str
    sheets_created: int

    def __str__(self) -> str:
        return f"{{'status': '{self.status}', 'file_name': '{self.file_name}', 'file_path': '{self.file_path}', 'sheets_created': {self.sheets_created}}}"


class EditSpreadsheetResponse(BaseModel):
    """Response for editing a spreadsheet."""

    model_config = ConfigDict(extra="forbid")

    status: str
    file_path: str
    operations_applied: int

    def __str__(self) -> str:
        return f"{{'status': '{self.status}', 'file_path': '{self.file_path}', 'operations_applied': {self.operations_applied}}}"


class AddTabResponse(BaseModel):
    """Response for adding a tab to a spreadsheet."""

    model_config = ConfigDict(extra="forbid")

    status: str
    tab_name: str
    file_path: str
    rows_added: int | None = None

    def __str__(self) -> str:
        base = f"{{'status': '{self.status}', 'tab_name': '{self.tab_name}', 'file_path': '{self.file_path}'"
        if self.rows_added is not None:
            base += f", 'rows_added': {self.rows_added}"
        base += "}"
        return base


class DeleteTabResponse(BaseModel):
    """Response for deleting a tab from a spreadsheet."""

    model_config = ConfigDict(extra="forbid")

    status: str
    tab_name: str
    tab_index: int
    file_path: str

    def __str__(self) -> str:
        return f"{{'status': '{self.status}', 'tab_name': '{self.tab_name}', 'tab_index': {self.tab_index}, 'file_path': '{self.file_path}'}}"


class DeleteSpreadsheetResponse(BaseModel):
    """Response for deleting a spreadsheet."""

    model_config = ConfigDict(extra="forbid")

    status: str
    file_path: str

    def __str__(self) -> str:
        return f"{{'status': '{self.status}', 'file_path': '{self.file_path}'}}"


class AddContentTextResponse(BaseModel):
    """Response for adding content to a cell."""

    model_config = ConfigDict(extra="forbid")

    status: str
    cell: str
    tab_index: int
    file_path: str

    def __str__(self) -> str:
        return f"{{'status': '{self.status}', 'cell': '{self.cell}', 'tab_index': {self.tab_index}, 'file_path': '{self.file_path}'}}"


class DeleteContentCellResponse(BaseModel):
    """Response for deleting content from a cell."""

    model_config = ConfigDict(extra="forbid")

    status: str
    cell: str
    tab_index: int
    file_path: str
    old_value: Any | None = None

    def __str__(self) -> str:
        base = f"{{'status': '{self.status}', 'cell': '{self.cell}', 'tab_index': {self.tab_index}, 'file_path': '{self.file_path}'"
        if self.old_value is not None:
            base += f", 'old_value': {repr(self.old_value)}"
        base += "}"
        return base


class ReadCsvResponse(BaseModel):
    """Response for reading a CSV file."""

    model_config = ConfigDict(extra="forbid")

    file_path: str
    headers: list[str] | None = None
    values: list[list[Any]]
    row_count: int
    column_count: int

    def __str__(self) -> str:
        if not self.values and not self.headers:
            return f"File: {self.file_path}\nTable: (empty)"

        table = _format_table_output(
            num_cols=self.column_count,
            values=self.values,
            headers=self.headers,
        )
        return f"File: {self.file_path}\nRows: {self.row_count}, Columns: {self.column_count}\nTable:\n{table}"


class FilterTabResponse(BaseModel):
    """Response for filtering a worksheet tab."""

    model_config = ConfigDict(extra="forbid")

    range: str
    filters_applied: int
    rows_matched: int
    total_rows: int
    values: list[list[Any]]
    headers: list[str] | None = None
    formulas: dict[str, str] | None = None
    diagnostic: str | None = None

    def __str__(self) -> str:
        if not self.values:
            base = (
                f"Range: {self.range}\n"
                f"Filters applied: {self.filters_applied}\n"
                f"Rows matched: {self.rows_matched}/{self.total_rows}\n"
                f"Table: (no matching rows)"
            )
            if self.diagnostic:
                base += f"\n\n{self.diagnostic}"
            return base

        num_cols = max(len(row) for row in self.values) if self.values else 0
        if self.headers:
            num_cols = max(num_cols, len(self.headers))

        table = _format_table_output(
            num_cols=num_cols,
            values=self.values,
            headers=self.headers,
            formulas=self.formulas,
        )
        return (
            f"Range: {self.range}\n"
            f"Filters applied: {self.filters_applied}\n"
            f"Rows matched: {self.rows_matched}/{self.total_rows}\n"
            f"Table:\n{table}"
        )
