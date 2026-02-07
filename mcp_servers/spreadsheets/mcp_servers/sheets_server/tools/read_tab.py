import os
from io import BytesIO
from typing import Annotated

from loguru import logger
from models.response import ReadTabRangeResponse, ReadTabSingleCellResponse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import Field
from utils.decorators import make_async_background
from utils.helpers import trim_empty_trailing_data
from utils.path_utils import PathTraversalError, resolve_under_root

# File size threshold for warning (3GB) - files above this will log a warning but still be processed
LARGE_FILE_WARNING_BYTES = 3 * 1024 * 1024 * 1024


@make_async_background
def read_tab(
    file_path: Annotated[str, Field(description="Path to the .xlsx file")],
    tab_index: Annotated[int, Field(description="0-based worksheet tab index", ge=0)],
    cell_range: Annotated[
        str | None, Field(description="Cell range like 'A1' or 'A1:C5'")
    ] = None,
) -> str:
    """Read a specific worksheet tab from a spreadsheet, optionally filtering by cell range."""

    if not isinstance(file_path, str) or not file_path:
        return "File path is required"
    if not file_path.startswith("/"):
        return "File path must start with /"
    if not file_path.lower().endswith(".xlsx"):
        return "File path must end with .xlsx"

    if not isinstance(tab_index, int) or tab_index < 0:
        return "Tab index must be a non-negative integer"

    try:
        target_path = resolve_under_root(file_path)
    except PathTraversalError:
        return f"Invalid path: {file_path}"

    try:
        if not os.path.exists(target_path):
            return f"File not found: {file_path}"
        if not os.path.isfile(target_path):
            return f"Not a file: {file_path}"

        # Log warning for very large files but still process them
        file_size = os.path.getsize(target_path)
        if file_size > LARGE_FILE_WARNING_BYTES:
            size_gb = file_size / (1024 * 1024 * 1024)
            logger.warning(
                f"Processing large file: {file_path} ({size_gb:.2f}GB). "
                "This may take longer and use significant memory."
            )
    except Exception as exc:
        return f"Failed to access file: {repr(exc)}"

    # Read file bytes once, then parse twice (values and formulas)
    # This reduces disk I/O compared to loading twice from disk
    try:
        with open(target_path, "rb") as f:
            file_bytes = f.read()
    except Exception as exc:
        return f"Failed to read file: {repr(exc)}"

    try:
        workbook_values = load_workbook(
            BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        return f"Failed to load workbook: {repr(exc)}"

    try:
        if tab_index >= len(workbook_values.sheetnames):
            sheet_count = len(workbook_values.sheetnames)
            workbook_values.close()
            return f"Tab index {tab_index} is out of range. Available sheets: {sheet_count}"

        worksheet_values = workbook_values[workbook_values.sheetnames[tab_index]]

        try:
            workbook_formulas = load_workbook(
                BytesIO(file_bytes), read_only=True, data_only=False
            )
        except Exception as exc:
            workbook_values.close()
            return f"Failed to load workbook for formulas: {repr(exc)}"

        worksheet_formulas = workbook_formulas[workbook_formulas.sheetnames[tab_index]]

        if cell_range is None:
            values = []
            formulas_dict = {}

            for row_idx, (row_values, row_formulas) in enumerate(
                zip(
                    worksheet_values.iter_rows(values_only=True),
                    worksheet_formulas.iter_rows(),
                    strict=True,
                ),
                start=1,
            ):
                values.append(list(row_values))
                for col_idx, cell in enumerate(row_formulas, start=1):
                    cell_value = cell.value
                    if isinstance(cell_value, str) and cell_value.startswith("="):
                        cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                        formulas_dict[cell_ref] = cell_value

            values = trim_empty_trailing_data(values)

            workbook_values.close()
            workbook_formulas.close()
            response = ReadTabRangeResponse(
                range="all",
                values=values,
                formulas=formulas_dict if formulas_dict else None,
            )
            return str(response)

        cell_range = cell_range.strip().upper()

        if ":" in cell_range:
            try:
                cell_obj_values = worksheet_values[cell_range]
                cell_obj_formulas = worksheet_formulas[cell_range]

                values = []
                formulas_dict = {}

                if not isinstance(cell_obj_values, tuple):
                    cell_obj_values = (cell_obj_values,)
                    cell_obj_formulas = (cell_obj_formulas,)

                for row_values, row_formulas in zip(
                    cell_obj_values, cell_obj_formulas, strict=True
                ):
                    if isinstance(row_values, tuple):
                        values.append([cell.value for cell in row_values])
                        for cell in row_formulas:
                            cell_value = cell.value
                            if isinstance(cell_value, str) and cell_value.startswith(
                                "="
                            ):
                                formulas_dict[cell.coordinate] = cell_value
                    else:
                        values.append([row_values.value])
                        cell_value = row_formulas.value
                        if isinstance(cell_value, str) and cell_value.startswith("="):
                            formulas_dict[row_formulas.coordinate] = cell_value

                workbook_values.close()
                workbook_formulas.close()
                response = ReadTabRangeResponse(
                    range=cell_range,
                    values=values,
                    formulas=formulas_dict if formulas_dict else None,
                )
                return str(response)
            except Exception as exc:
                workbook_values.close()
                workbook_formulas.close()
                return f"Invalid cell range '{cell_range}': {repr(exc)}"
        else:
            try:
                cell_value = worksheet_values[cell_range].value
                cell_formula_value = worksheet_formulas[cell_range].value

                formula = None
                if isinstance(
                    cell_formula_value, str
                ) and cell_formula_value.startswith("="):
                    formula = cell_formula_value

                workbook_values.close()
                workbook_formulas.close()
                response = ReadTabSingleCellResponse(
                    cell=cell_range, value=cell_value, formula=formula
                )
                return str(response)
            except Exception as exc:
                workbook_values.close()
                workbook_formulas.close()
                return f"Invalid cell reference '{cell_range}': {repr(exc)}"

    except Exception as exc:
        try:
            workbook_values.close()
        except Exception:
            pass
        return f"Unexpected error: {repr(exc)}"
