"""Filter data from a spreadsheet worksheet based on conditions."""

import os
import re
from io import BytesIO
from typing import Annotated, Any

from loguru import logger
from models.response import FilterTabResponse
from models.sheet import FilterCondition
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from pydantic import Field, TypeAdapter
from utils.decorators import make_async_background
from utils.helpers import trim_empty_trailing_data
from utils.path_utils import PathTraversalError, resolve_under_root

# File size threshold for warning (3GB)
LARGE_FILE_WARNING_BYTES = 3 * 1024 * 1024 * 1024

# Regex pattern for column letters
_COLUMN_LETTER_PATTERN = re.compile(r"^[A-Za-z]+$")

# TypeAdapter for validating filter conditions
_filter_condition_adapter = TypeAdapter(list[FilterCondition])


def _is_column_letter(value: str) -> bool:
    """Check if a string is a valid Spreadsheets column letter (A, B, AA, etc.)."""
    return bool(_COLUMN_LETTER_PATTERN.match(value))


def _resolve_column_index(
    column: str, headers: list[str] | None, use_headers: bool, num_cols: int
) -> tuple[int | None, str | None]:
    """Resolve a column reference to a 0-based column index.

    Args:
        column: Column letter (A, B) or header name
        headers: List of header values (if use_headers is True)
        use_headers: Whether to allow header name lookups
        num_cols: Number of columns in the data (for error messages)

    Returns:
        Tuple of (0-based column index or None, error message or None)
    """
    # When use_headers is enabled, try header name lookup FIRST
    # This prevents short header names like "ID", "Qty", "Amt" from being
    # misinterpreted as Spreadsheets column letters (ID=238, etc.)
    if use_headers and headers:
        # Case-insensitive header matching
        column_lower = column.lower()
        for idx, header in enumerate(headers):
            if header is not None and str(header).lower() == column_lower:
                return idx, None

    # Fall back to column letter interpretation
    if _is_column_letter(column):
        try:
            col_idx = column_index_from_string(column.upper()) - 1
            if col_idx < num_cols:
                return col_idx, None
            else:
                available_cols = ", ".join(
                    get_column_letter(i + 1) for i in range(min(num_cols, 10))
                )
                if num_cols > 10:
                    available_cols += f"... ({num_cols} columns total)"
                return (
                    None,
                    f"Column '{column}' is out of range. Available columns: {available_cols}",
                )
        except ValueError:
            pass

    # Column not found - provide helpful error
    if use_headers and headers:
        # Show available headers
        available_headers = ", ".join(f"'{h}'" for h in headers[:10] if h)
        if len(headers) > 10:
            available_headers += f"... ({len(headers)} headers total)"
        available_cols = ", ".join(
            get_column_letter(i + 1) for i in range(min(num_cols, 10))
        )
        return None, (
            f"Column '{column}' not found. "
            f"Available headers: {available_headers}. "
            f"Or use column letters: {available_cols}"
        )
    else:
        available_cols = ", ".join(
            get_column_letter(i + 1) for i in range(min(num_cols, 10))
        )
        if num_cols > 10:
            available_cols += f"... ({num_cols} columns total)"
        return None, f"Column '{column}' not found. Available columns: {available_cols}"


def _evaluate_condition(cell_value: Any, operator: str, filter_value: Any) -> bool:
    """Evaluate a single filter condition against a cell value.

    Args:
        cell_value: The value from the cell
        operator: The filter operator
        filter_value: The value to compare against

    Returns:
        True if the condition is satisfied, False otherwise
    """
    # Handle is_empty and is_not_empty first
    if operator == "is_empty":
        return cell_value is None or cell_value == ""
    if operator == "is_not_empty":
        return cell_value is not None and cell_value != ""

    # For other operators, we need a filter_value
    if filter_value is None:
        return False

    # String operations
    if operator == "contains":
        if cell_value is None:
            return False
        return str(filter_value).lower() in str(cell_value).lower()

    if operator == "not_contains":
        if cell_value is None:
            return True
        return str(filter_value).lower() not in str(cell_value).lower()

    if operator == "starts_with":
        if cell_value is None:
            return False
        return str(cell_value).lower().startswith(str(filter_value).lower())

    if operator == "ends_with":
        if cell_value is None:
            return False
        return str(cell_value).lower().endswith(str(filter_value).lower())

    # Equality operations
    if operator == "equals":
        if cell_value is None and filter_value is None:
            return True
        if cell_value is None or filter_value is None:
            return False
        # Try numeric comparison first
        try:
            return float(cell_value) == float(filter_value)
        except (ValueError, TypeError):
            pass
        # Fall back to string comparison (case-insensitive)
        return str(cell_value).lower() == str(filter_value).lower()

    if operator == "not_equals":
        if cell_value is None and filter_value is None:
            return False
        if cell_value is None or filter_value is None:
            return True
        # Try numeric comparison first
        try:
            return float(cell_value) != float(filter_value)
        except (ValueError, TypeError):
            pass
        # Fall back to string comparison (case-insensitive)
        return str(cell_value).lower() != str(filter_value).lower()

    # Numeric comparisons
    if operator in (
        "greater_than",
        "less_than",
        "greater_than_or_equal",
        "less_than_or_equal",
    ):
        if cell_value is None:
            return False
        try:
            cell_num = float(cell_value)
            filter_num = float(filter_value)
        except (ValueError, TypeError):
            return False

        if operator == "greater_than":
            return cell_num > filter_num
        if operator == "less_than":
            return cell_num < filter_num
        if operator == "greater_than_or_equal":
            return cell_num >= filter_num
        if operator == "less_than_or_equal":
            return cell_num <= filter_num

    return False


def _apply_filters(
    row: list[Any],
    conditions: list[FilterCondition],
    column_indices: dict[str, int],
    match_all: bool,
) -> bool:
    """Apply filter conditions to a row using pre-computed column indices.

    Args:
        row: The data row to filter
        conditions: List of filter conditions
        column_indices: Pre-computed mapping of column names to 0-based indices
        match_all: If True, all conditions must match (AND); if False, any match (OR)

    Returns:
        True if the row passes the filter, False otherwise
    """
    if not conditions:
        return True

    results = []
    for condition in conditions:
        col_idx = column_indices.get(condition.column)
        if col_idx is None:
            # Column not found - condition fails (should not happen after validation)
            results.append(False)
            continue

        cell_value = row[col_idx] if col_idx < len(row) else None
        result = _evaluate_condition(cell_value, condition.operator, condition.value)
        results.append(result)

    if match_all:
        return all(results)
    else:
        return any(results)


@make_async_background
def filter_tab(
    file_path: Annotated[str, Field(description="Path to the .xlsx file")],
    tab_index: Annotated[int, Field(description="0-based worksheet tab index", ge=0)],
    filters: Annotated[
        list[dict[str, Any]],
        Field(
            description=(
                "List of filter conditions. Each condition has: "
                "'column' (column letter like 'A' or header name), "
                "'operator' (equals, not_equals, greater_than, less_than, "
                "greater_than_or_equal, less_than_or_equal, contains, not_contains, "
                "starts_with, ends_with, is_empty, is_not_empty), "
                "'value' (value to compare, optional for is_empty/is_not_empty)"
            )
        ),
    ],
    cell_range: Annotated[
        str | None,
        Field(description="Optional cell range to filter within, e.g., 'A1:D100'"),
    ] = None,
    match_all: Annotated[
        bool,
        Field(
            description=(
                "If True (default), all conditions must match (AND logic); "
                "if False, any condition matching is sufficient (OR logic)"
            )
        ),
    ] = True,
    use_headers: Annotated[
        bool,
        Field(
            description=(
                "If True (default), first row is treated as headers and "
                "column names can reference header values"
            )
        ),
    ] = True,
) -> str:
    """Filter data from a worksheet tab based on conditions.

    Supports filtering by column values with various operators (equals, contains,
    greater_than, etc.). Columns can be referenced by letter (A, B) or by header
    name if use_headers is True.
    """
    # Validate file_path
    if not isinstance(file_path, str) or not file_path:
        return "File path is required"
    if not file_path.startswith("/"):
        return "File path must start with /"
    if not file_path.lower().endswith(".xlsx"):
        return "File path must end with .xlsx"

    # Validate tab_index
    if not isinstance(tab_index, int) or tab_index < 0:
        return "Tab index must be a non-negative integer"

    # Validate and parse filters
    if not isinstance(filters, list):
        return "Filters must be a list"
    if not filters:
        return "At least one filter condition is required"

    try:
        conditions = _filter_condition_adapter.validate_python(filters)
    except Exception as exc:
        return f"Invalid filter conditions: {repr(exc)}"

    # Resolve path
    try:
        target_path = resolve_under_root(file_path)
    except PathTraversalError:
        return f"Invalid path: {file_path}"

    # Check file exists
    try:
        if not os.path.exists(target_path):
            return f"File not found: {file_path}"
        if not os.path.isfile(target_path):
            return f"Not a file: {file_path}"

        file_size = os.path.getsize(target_path)
        if file_size > LARGE_FILE_WARNING_BYTES:
            size_gb = file_size / (1024 * 1024 * 1024)
            logger.warning(
                f"Processing large file: {file_path} ({size_gb:.2f}GB). "
                "This may take longer and use significant memory."
            )
    except Exception as exc:
        return f"Failed to access file: {repr(exc)}"

    # Read file bytes
    try:
        with open(target_path, "rb") as f:
            file_bytes = f.read()
    except Exception as exc:
        return f"Failed to read file: {repr(exc)}"

    # Load workbook
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        return f"Failed to load workbook: {repr(exc)}"

    try:
        # Validate tab index
        if tab_index >= len(workbook.sheetnames):
            sheet_count = len(workbook.sheetnames)
            workbook.close()
            return f"Tab index {tab_index} is out of range. Available sheets: {sheet_count}"

        worksheet = workbook[workbook.sheetnames[tab_index]]

        # Read data based on cell_range
        all_rows: list[list[Any]] = []
        range_str = "all"

        if cell_range is not None:
            cell_range = cell_range.strip().upper()
            range_str = cell_range

            if ":" not in cell_range:
                workbook.close()
                return "Cell range must be a range like 'A1:D100', not a single cell"

            try:
                cell_obj = worksheet[cell_range]
                if not isinstance(cell_obj, tuple):
                    cell_obj = (cell_obj,)

                for row in cell_obj:
                    if isinstance(row, tuple):
                        all_rows.append([cell.value for cell in row])
                    else:
                        all_rows.append([row.value])
            except Exception as exc:
                workbook.close()
                return f"Invalid cell range '{cell_range}': {repr(exc)}"
        else:
            # Read entire sheet
            for row in worksheet.iter_rows(values_only=True):
                all_rows.append(list(row))

        workbook.close()

        # Trim empty trailing data
        all_rows = trim_empty_trailing_data(all_rows)

        if not all_rows:
            return str(
                FilterTabResponse(
                    range=range_str,
                    filters_applied=len(conditions),
                    rows_matched=0,
                    total_rows=0,
                    values=[],
                    headers=None,
                )
            )

        # Extract headers if use_headers is True
        headers: list[str] | None = None
        data_rows = all_rows

        if use_headers and all_rows:
            headers = [str(v) if v is not None else "" for v in all_rows[0]]
            data_rows = all_rows[1:]

        # Determine number of columns for validation
        num_cols = max(len(row) for row in all_rows) if all_rows else 0

        # First, validate all column references before filtering
        column_errors: list[str] = []
        column_indices: dict[str, int] = {}
        for condition in conditions:
            if condition.column not in column_indices:
                col_idx, error = _resolve_column_index(
                    condition.column, headers, use_headers, num_cols
                )
                if col_idx is not None:
                    column_indices[condition.column] = col_idx
                elif error and error not in column_errors:
                    column_errors.append(error)

        # If there are column resolution errors, return them immediately
        if column_errors:
            error_msg = "Filter error: " + "; ".join(column_errors)
            return error_msg

        # Apply filters to each row
        matched_rows: list[list[Any]] = []
        for row in data_rows:
            if _apply_filters(row, conditions, column_indices, match_all):
                matched_rows.append(row)

        # Build diagnostic info if no matches found
        diagnostic: str | None = None
        if len(matched_rows) == 0 and data_rows:
            # Collect sample values from filtered columns to help debug
            diag_parts = []
            for condition in conditions:
                col_idx = column_indices.get(condition.column)
                if col_idx is not None:
                    # Get sample values from this column (first 5 non-empty)
                    sample_values = []
                    for row in data_rows[:10]:
                        if col_idx < len(row) and row[col_idx] is not None:
                            val = row[col_idx]
                            val_type = type(val).__name__
                            sample_values.append(f"{repr(val)} ({val_type})")
                            if len(sample_values) >= 5:
                                break
                    if sample_values:
                        col_name = condition.column
                        if headers and col_idx < len(headers):
                            col_name = f"{headers[col_idx]} (column {get_column_letter(col_idx + 1)})"
                        diag_parts.append(
                            f"Column '{col_name}' sample values: {', '.join(sample_values)}"
                        )
                    else:
                        diag_parts.append(
                            f"Column '{condition.column}' has no values in first 10 rows"
                        )
            if diag_parts:
                diagnostic = (
                    "No rows matched. Debug info:\n"
                    + "\n".join(f"  - {p}" for p in diag_parts)
                    + f"\nFilter attempted: {', '.join(f'{c.column} {c.operator} {repr(c.value)}' for c in conditions)}"
                )

        response = FilterTabResponse(
            range=range_str,
            filters_applied=len(conditions),
            rows_matched=len(matched_rows),
            total_rows=len(data_rows),
            values=matched_rows,
            headers=headers,
            diagnostic=diagnostic,
        )
        return str(response)

    except Exception as exc:
        try:
            workbook.close()
        except Exception:
            pass
        return f"Unexpected error: {repr(exc)}"
